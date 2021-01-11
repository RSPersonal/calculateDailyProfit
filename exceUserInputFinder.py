#! Python 3
# Excel Cell finder - A program to find cells based on user input data

# TO DO:
# 1. Get the right libraries for this script 
# 2. Get user input as [keyphrase] and store this in list
# 3. Script opens excel file and searches on all sheets for keyphrases
# 4. If keyphrase is found highlight this in the excel file
# 5. (Optional) Script creates a empty sheet and takes the keyphrases + values and writes in a overview on the newly created sheet

#importing dependencies
import openpyxl as op
import os.path
import sys

#Specifying cell color
my_red = op.styles.colors.Color(rgb='00FF0000')
my_fill = op.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

#Asking input file from user. The user does not need to write the .xlsx format. The script does this automatically
user_input_file = input("Please specify input file name, the file also needs to be in the current directory: ")
user_input_file = user_input_file + ".xlsx"

#checking if file is in directory 
while True:
    if os.path.isfile(user_input_file):
        print("\nFile found in directory!")
        break
    else:
        print("\n File is not found, please place the file in the current directory!")
        break
        exit()

#Loading in Excel workbook and make working sheet -> active worksheet 
#To-Do: Make script look into all sheets instead of active 
wb1 = op.load_workbook(user_input_file)
active_worksheet = wb1.active 

#Creating empty search list for user input keyphrases. These keyphrases are going to searched in the excel file. 
search_list = [["Marktwaarde"], ["k.k"], ["kapitaalcorrecties"], ["huurwaarde"], ["totale huurwaarde"], ['correcties'], ['huuropbrengsten']]
keywords_found = []

#Asking user for keywords and appending them to the search list. User can enter as much as he wants.
print("\nPlease enter your keyword, you can enter as many keywords as you want.")
while True:
    user_input = input("\nKeyword: ")
    search_list.append([user_input])
    print(user_input + " added to searchlist.")
    yes_or_no = input("\n Do you want to enter another keyword? ('Yes/No')")
    if yes_or_no == 'Yes':
        continue
    elif yes_or_no == 'No':
        break
    else:
        while True:
            yes_or_no = input("\nWrong input! Please enter Yes or No!")
            if yes_or_no == 'Yes':
                break
            elif yes_or_no == 'No':
                break
    break

#Searching in all rows for keyphrases and print out if found or not!
for list in search_list:
    for element in list:
        for row in active_worksheet.iter_rows(min_col=1, max_col=50, min_row=1, max_row=160):
            for any_cell in row:
                if any_cell.value == element:
                    print("\n Keyword: " + element + " found in sheet!")
                    keywords_found.append(element)
                    any_cell.fill = my_fill
                    
print("\nKeywords found =" + str(keywords_found) + " Highlighted the keywords in Excel file") 
print("\nKeywords not found =" + str(search_list))

exit = input("\nDo you wish to exit the program? (Y = Yes, N = No)")
if exit == "Y":
    exit()
else:
    continue
